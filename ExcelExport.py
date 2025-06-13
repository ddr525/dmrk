import datetime
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.marker import Marker
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl import Workbook

from utilities import toFixed


def export_to_excel(heating_data, exp_num):
    #print(heating_data["data"])
    graph = heating_data["График"]
    # print(graph)
    t = np.array(graph["t"])
    time_H = float(graph["время"])  # Преобразуем в float для надежности
    twG = np.array(graph["Печь (верх)"])
    twnG = np.array(graph["Печь (низ)"])
    tmvG = np.array(graph["Металл (верх)"])
    tmcG = np.array(graph["Металл (центр)"])
    tmnG = np.array(graph["Металл (низ)"])
    tAvg = heating_data["Расчет нагрева металла"]["Среднемассовая температура металла, °C"]
    tP = heating_data["Расчет нагрева металла"]["Конечный перепад температур, °C"]
    frostmourne_hungers = heating_data["Расчет нагрева металла"]["Температура раската за пятой клетью (T5), °C"]


    wb = Workbook()
    ws = wb.active
    ws['A1'] = f"Опыт № {exp_num}"
    ws.merge_cells('B1:J1')
    ws['B1'] = 'Расчет нагрева сляба в методической печи'
    ws['B1'].font = Font(bold=True)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A2:J2')
    ws['A2'] = 'температурный режим в соответсвии с действующей ТИ при времени нагрева 4ч'
    ws['A2'].font = Font(bold=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ["Зона", "Время нагрева, мин", "Печь (верх)", "Печь (низ)", "Металл (верх)", "Металл (центр)", "Металл (низ)"]
    zones = ["посад", "метод. зона", "1/2 зона", "3/4 зона", "5 зона"]
    ws.append(headers)

    percentages = [0]
    prev = 0
    for e in heating_data["data"]:
        if(e == "Время нагрева (time_H)"):
            continue
        if heating_data["data"][e] < 1:
            prev = prev + heating_data["data"][e] * 100
        else:
            prev = prev + heating_data["data"][e] 
        percentages.append(prev )
    # print(percentages)
    indices = [np.abs(t - (time_H * p / 100)).argmin() for p in percentages]

    row = [
            zones[0],
            # t[idx],
            0,
            # time_H,  # same time_H value for every row
            heating_data["t_data"]["Температура верха печи при посаде"],
            heating_data["t_data"]["Температура низа печи при посаде"],
            heating_data["t_data"]["Температура сляба при посаде"],
            heating_data["t_data"]["Температура сляба при посаде"],
            heating_data["t_data"]["Температура сляба при посаде"],
        ]
    ws.append(row)

    for i in range(1, len(indices)):
        row = [
            zones[i],
            # t[idx],
            (time_H * percentages[i])/100,
            # time_H,  # same time_H value for every row
            twG[indices[i]],
            twnG[indices[i]],
            tmvG[indices[i]],
            tmcG[indices[i]],
            tmnG[indices[i]],
        ]
        ws.append(row)

    

    #heating_data["Расчет нагрева металла"]["Методическая зона"]["tпечи, °C"]
    #heating_data["Расчет нагрева металла"]["Методическая зона"]["tниз, °C"]

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 14

    chart = LineChart()
    data_ref = Reference(ws, min_col=3, min_row=3, max_col=7, max_row=len(indices)+3)
    categories_ref = Reference(ws, min_col=2, min_row=4, max_row=len(indices)+3)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(categories_ref)
    chart_position = f"A{len(indices) + 6}" 

    chart.title = "Распределение температур по длине МП"
    chart.legend.position = 'r'
    chart.layout = Layout(
        ManualLayout(
        x=0.05, y=0.1,
        h=0.8, w=0.75,
        xMode="edge",
        yMode="edge",
        )
    )
    chart.x_axis.title = 'Время нагрева, мин'
    chart.y_axis.title = 'Температура °C'
    chart.x_axis.visible = True  # Ensure x-axis is visible
    chart.y_axis.visible = True  # Ensure y-axis is visible
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    
    ws.add_chart(chart, chart_position)
    chart.width = 24  # Width of the chart
    chart.height = 10
    

    for row in ws[f'A1:J{len(indices) + 3}']:
        for cell in row:
            cell.border = Border(left=Side(border_style="thick", color="000000"),right=Side(border_style="thick", color="000000"),top=Side(border_style="thick", color="000000"),bottom=Side(border_style="thick", color="000000"))
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for row in ws[f'B4:B8']:
        for cell in row:
            cell.number_format = "0.00"

    for row in ws[f'C4:G{len(indices) + 3}']:
        for cell in row:
            cell.border = Border(left=Side(border_style="thin", color="000000"),right=Side(border_style="thin", color="000000"),top=Side(border_style="thin", color="000000"),bottom=Side(border_style="thin", color="000000"))
            cell.number_format = "0.0"
    for row in ws[f'B1:J1']:
        for cell in row:
            cell.border = Border(left=Side(border_style="thin", color="000000"),right=Side(border_style="thin", color="000000"),top=Side(border_style="thin", color="000000"),)
    for row in ws[f'A2:J2']:
        for cell in row:
            cell.border = Border(left=Side(border_style="thin", color="000000"),right=Side(border_style="thin", color="000000"),bottom=Side(border_style="thin", color="000000"),)
    for row in ws[f'C8:G8']:
        for cell in row:
            cell.border = Border(left=Side(border_style="thin", color="000000"),right=Side(border_style="thin", color="000000"),top=Side(border_style="thin", color="000000"),bottom=Side(border_style="thick", color="000000"))

    ws.merge_cells('H3:H7')
    ws['H3'] = "Среднемассовая температура металла"
    ws['H3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['H8'] = float(tAvg)
    ws['H8'].number_format = "0.0"
    ws.merge_cells('I3:I7')
    ws['I3'] = "Конечный перепад температур"
    ws['I3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['I8'] = float(tP)
    ws['I8'].number_format = "0.0"
    ws.merge_cells('J3:J7')
    ws['J3'] = "Температура раската за пятой клетью (T5), °C"
    ws['J3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['J8'] = float(frostmourne_hungers)
    ws['J8'].number_format = "0.0"

    markers = [
        Marker(symbol='circle', size=6),
        Marker(symbol='square', size=6),
        Marker(symbol='diamond', size=6),
        Marker(symbol='x', size=6),
        Marker(symbol='star', size=6)
    ]

    for i, series in enumerate(chart.series):
        series.marker = markers[i % len(markers)]

    current_time = datetime.datetime.now()
    filename = current_time.strftime("%Y-%m-%d_%H-%M-%S") + ".xlsx"
    # filename = 'export.xlsx'
    wb.save(filename)
    return filename
